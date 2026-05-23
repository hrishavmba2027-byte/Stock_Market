"""Feature Engineering for NSE Stock Data.

This module provides two distinct modes of operation:

1.  **Inline import (production workflow)**
    ``compute_indicators(df)`` is imported directly by ``main.py``,
    ``monthly_finetune.py``, and other modules.  The function receives a
    single-symbol DataFrame already loaded from Google Sheets (or another
    source), computes 29 technical indicators, attaches 30 forward-return
    labels, and returns the enriched DataFrame.  It never touches the
    filesystem.

2.  **Standalone batch script (offline / training data refresh)**
    ``python Feature_Engineering.py`` reads
    ``Data/nse_stock_data_train.xlsx`` (the 51 MB training corpus), runs
    ``compute_indicators`` on every sheet, and writes the enriched workbook
    back to the same file.  This is a manual, offline step — it is NOT
    invoked by the Docker workflow.

Design invariants
-----------------
* ``compute_indicators`` MUST NOT import from ``main.py`` or any app module
  (to avoid circular imports).
* Indicator columns are lagged by one step (``shift(1)``) to prevent
  look-ahead bias: the feature row at time *t* reflects information up to
  *t-1*.
* Forward-return labels (``y_logret_h1`` … ``y_logret_h30``) and the
  ``has_labels`` flag are targets, NOT model features.  They are stripped by
  callers before building the feature matrix.
* A ``FeatureEngineeringError`` is raised instead of silently returning the
  unmodified DataFrame when OHLCV columns cannot be located.  Callers are
  expected to catch this and surface it as a skipped-symbol warning.
"""

from __future__ import annotations

import logging
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings_module_available = False
try:
    import warnings
    warnings.filterwarnings("ignore")
    warnings_module_available = True
except Exception:
    pass

_log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Public error type
# ---------------------------------------------------------------------------

class FeatureEngineeringError(ValueError):
    """Raised when feature engineering cannot proceed for a symbol.

    Replaces the old silent-return path so callers get an explicit signal
    instead of an unmodified DataFrame that looks valid but has no indicators.
    """


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = {"open", "high", "low", "close", "volume"}
DATE_COLUMN_VARIANTS = ["date", "Date", "DATE", "datetime", "Datetime", "time", "Time"]

# Forward-return label settings (P0.1)
FORWARD_HORIZONS: Tuple[int, ...] = tuple(range(1, 31))   # h = 1, 2, …, 30
FORWARD_LABEL_PREFIX = "y_logret_h"
HAS_LABELS_COL = "has_labels"

# Columns produced by compute_indicators that the model uses (29 total).
# MFI_14 is intentionally excluded — it was computed historically but was never
# added to pipeline_metadata.json feature_columns, so including it here
# prevents a silent divergence between what FE produces and what the model reads.
INDICATOR_COLUMNS: Tuple[str, ...] = (
    "RSI_14",
    "MACD_12_26", "MACD_Signal_9", "MACD_Histogram",
    "Stochastic_%K", "Stochastic_%D",
    "SMA_5", "SMA_20", "SMA_50",
    "EMA_12", "EMA_26", "EMA_50",
    "ADX_14",
    "BB_Upper_20", "BB_Middle_20", "BB_Lower_20",
    "ATR_14",
    "OBV", "VWAP",
    "Daily_Return_%", "Log_Return_%",
    "ROC_12", "CCI_20", "Williams_%R",
)  # len = 24; plus 5 OHLCV = 29 total

# Decision-layer categorical features (Google Sheets only — never model inputs)
DECISION_FEATURE_COLUMNS = [
    "RSI_7", "SMA_5_20", "SMA_20_50", "EMA_12_26", "EMA_26_50",
    "BB_indicator", "BB_trend_state", "BB_volatility_state",
    "BB_overbought_oversold", "RSI_indicator_7", "RSI_indicator_14",
]

RSI_BIN_MIN_HISTORY = 30
_RSI_FALLBACK_THRESHOLDS = (30.0, 50.0, 70.0)


# ===========================================================================
# Date / OHLCV detection helpers
# ===========================================================================

def detect_date_column(df: pd.DataFrame) -> Optional[str]:
    """Return the name of the date column, or None if not detected."""
    for col in DATE_COLUMN_VARIANTS:
        if col in df.columns:
            return col
    for col in df.columns:
        if "date" in str(col).lower() or "time" in str(col).lower():
            return col
    if df.columns.size > 0:
        first_col = df.columns[0]
        try:
            pd.to_datetime(df[first_col])
            return first_col
        except Exception:
            pass
    return None


def ensure_date_column(df: pd.DataFrame, date_col: str) -> pd.DataFrame:
    """Parse *date_col* to datetime, drop invalid rows, sort chronologically."""
    if not date_col:
        return df
    try:
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce", format="mixed")
    except TypeError:
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=[date_col])
    df = df.sort_values(by=date_col).reset_index(drop=True)
    return df


def find_ohlcv_columns(df: pd.DataFrame) -> Optional[Dict[str, str]]:
    """Locate OHLCV columns using case-insensitive matching.

    Returns a dict ``{"open": col, "high": col, …}`` with all five keys,
    or ``None`` if any of the five cannot be found.
    """
    col_lower = {col: str(col).lower() for col in df.columns}

    # Pass 1 — exact lowercase match
    ohlcv_map: Dict[str, str] = {}
    for name in ("open", "high", "low", "close", "volume"):
        matches = [col for col, lc in col_lower.items() if lc == name]
        if matches:
            ohlcv_map[name] = matches[0]
    if len(ohlcv_map) == 5:
        return ohlcv_map

    # Pass 2 — prefix pattern (e.g. Open_RELIANCE)
    ohlcv_map = {}
    for col, lc in col_lower.items():
        for prefix, key in (
            ("open_", "open"), ("high_", "high"), ("low_", "low"),
            ("close_", "close"), ("volume_", "volume"),
        ):
            if lc.startswith(prefix):
                ohlcv_map.setdefault(key, col)
    if len(ohlcv_map) == 5:
        return ohlcv_map

    return None


def find_close_column(df: pd.DataFrame) -> Optional[str]:
    """Locate the Close column leniently (exact, then Close_SYMBOL pattern)."""
    for col in df.columns:
        if str(col).strip().lower() == "close":
            return col
    for col in df.columns:
        if str(col).strip().lower().startswith("close_"):
            return col
    return None


# ===========================================================================
# Core technical indicator implementations
# ===========================================================================

def compute_rsi(close: pd.Series, period: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.where(delta > 0, 0.0).rolling(window=period).mean()
    loss = (-delta.where(delta < 0, 0.0)).rolling(window=period).mean()
    rs = np.where(loss != 0, gain / loss, np.nan)
    return pd.Series(100.0 - (100.0 / (1.0 + rs)), index=close.index)


def compute_macd(
    close: pd.Series, fast: int = 12, slow: int = 26, signal: int = 9
) -> Tuple[pd.Series, pd.Series, pd.Series]:
    ema_fast = close.ewm(span=fast, adjust=False).mean()
    ema_slow = close.ewm(span=slow, adjust=False).mean()
    macd = ema_fast - ema_slow
    sig = macd.ewm(span=signal, adjust=False).mean()
    return macd, sig, macd - sig


def compute_stochastic(
    high: pd.Series, low: pd.Series, close: pd.Series,
    k_period: int = 14, d_period: int = 3
) -> Tuple[pd.Series, pd.Series]:
    lo = low.rolling(k_period).min()
    hi = high.rolling(k_period).max()
    denom = hi - lo
    k = 100.0 * ((close - lo) / denom)
    return k, k.rolling(d_period).mean()


def compute_sma(close: pd.Series, period: int) -> pd.Series:
    return close.rolling(period).mean()


def compute_ema(close: pd.Series, period: int) -> pd.Series:
    return close.ewm(span=period, adjust=False).mean()


def compute_adx(
    high: pd.Series, low: pd.Series, close: pd.Series, period: int = 14
) -> pd.Series:
    try:
        plus_dm = high.diff().clip(lower=0.0)
        minus_dm = (-low.diff()).clip(lower=0.0)
        # Keep only directional moves (not both sides simultaneously)
        plus_dm = plus_dm.where(plus_dm > (-low.diff()).clip(lower=0.0), 0.0)
        minus_dm = minus_dm.where(minus_dm > high.diff().clip(lower=0.0), 0.0)

        tr = pd.concat([
            high - low,
            (high - close.shift(1)).abs(),
            (low - close.shift(1)).abs(),
        ], axis=1).max(axis=1)

        tr_s = tr.rolling(period).sum()
        plus_di = 100.0 * (plus_dm.rolling(period).sum() / tr_s)
        minus_di = 100.0 * (minus_dm.rolling(period).sum() / tr_s)
        di_sum = (plus_di + minus_di).abs()
        dx = 100.0 * ((plus_di - minus_di).abs() / di_sum.replace(0, np.nan))
        return dx.rolling(period).mean()
    except Exception as exc:
        _log.warning("ADX calculation failed: %s", exc)
        return pd.Series(np.nan, index=close.index)


def compute_bollinger_bands(
    close: pd.Series, period: int = 20, std_dev: float = 2.0
) -> Tuple[pd.Series, pd.Series, pd.Series]:
    mid = close.rolling(period).mean()
    std = close.rolling(period).std()
    return mid + std_dev * std, mid, mid - std_dev * std


def compute_atr(
    high: pd.Series, low: pd.Series, close: pd.Series, period: int = 14
) -> pd.Series:
    tr = pd.concat([
        high - low,
        (high - close.shift(1)).abs(),
        (low - close.shift(1)).abs(),
    ], axis=1).max(axis=1)
    return tr.rolling(period).mean()


def compute_obv(close: pd.Series, volume: pd.Series) -> pd.Series:
    price_diff = close.diff().fillna(0.0)
    return (np.sign(price_diff) * volume).cumsum()


def compute_vwap(
    high: pd.Series, low: pd.Series, close: pd.Series, volume: pd.Series
) -> pd.Series:
    try:
        tp = (high + low + close) / 3.0
        return (tp * volume).cumsum() / volume.cumsum()
    except Exception as exc:
        _log.warning("VWAP calculation failed: %s", exc)
        return pd.Series(np.nan, index=close.index)


def compute_daily_return(close: pd.Series) -> pd.Series:
    return close.pct_change() * 100.0


def compute_log_return(close: pd.Series) -> pd.Series:
    return np.log(close / close.shift(1)) * 100.0


def compute_roc(close: pd.Series, period: int = 12) -> pd.Series:
    return ((close - close.shift(period)) / close.shift(period)) * 100.0


def compute_cci(
    high: pd.Series, low: pd.Series, close: pd.Series, period: int = 20
) -> pd.Series:
    try:
        tp = (high + low + close) / 3.0
        sma = tp.rolling(period).mean()
        mad = tp.rolling(period).apply(lambda x: np.abs(x - x.mean()).mean(), raw=False)
        return (tp - sma) / (0.015 * mad)
    except Exception as exc:
        _log.warning("CCI calculation failed: %s", exc)
        return pd.Series(np.nan, index=close.index)


def compute_williams_r(
    high: pd.Series, low: pd.Series, close: pd.Series, period: int = 14
) -> pd.Series:
    hh = high.rolling(period).max()
    ll = low.rolling(period).min()
    return -100.0 * ((hh - close) / (hh - ll))


# ===========================================================================
# Forward-return label helpers (P0.1)
# ===========================================================================

def forward_label_columns(horizons: Tuple[int, ...] = FORWARD_HORIZONS) -> List[str]:
    return [f"{FORWARD_LABEL_PREFIX}{h}" for h in horizons]


def compute_forward_log_returns(
    close: pd.Series, horizons: Tuple[int, ...] = FORWARD_HORIZONS
) -> Dict[str, pd.Series]:
    """Compute log(Close[t+h] / Close[t]) for each horizon h."""
    log_close = np.log(close)
    return {
        f"{FORWARD_LABEL_PREFIX}{h}": log_close.shift(-h) - log_close
        for h in horizons
    }


def attach_forward_labels(
    df: pd.DataFrame,
    close_col: str = "close",
    horizons: Tuple[int, ...] = FORWARD_HORIZONS,
) -> pd.DataFrame:
    """Attach label columns + ``has_labels`` to *df*.  Does not shift."""
    df_out = df.copy()
    labels = compute_forward_log_returns(df_out[close_col], horizons=horizons)
    label_cols = list(labels.keys())
    for name, series in labels.items():
        df_out[name] = series.values
    df_out[HAS_LABELS_COL] = df_out[label_cols].notna().all(axis=1)
    return df_out


# ===========================================================================
# Decision-layer categorical features (Sheets-only, never model inputs)
# ===========================================================================

def _finite(series: pd.Series) -> pd.Series:
    return pd.to_numeric(pd.Series(series), errors="coerce").replace(
        [np.inf, -np.inf], np.nan
    )


def rsi_percentile_thresholds(
    rsi_values,
) -> Optional[Tuple[float, float, float]]:
    arr = _finite(pd.Series(list(rsi_values), dtype="float64")).dropna()
    if len(arr) < RSI_BIN_MIN_HISTORY:
        return None
    return (
        float(arr.quantile(0.25)),
        float(arr.quantile(0.50)),
        float(arr.quantile(0.75)),
    )


def classify_rsi_bins(rsi_series, thresholds) -> List:
    p25, p50, p75 = (
        thresholds if thresholds is not None else _RSI_FALLBACK_THRESHOLDS
    )
    out = []
    for v in _finite(rsi_series):
        if pd.isna(v):
            out.append("")
        elif v > p75:
            out.append("highly overbought")
        elif v > p50:
            out.append("slightly overbought")
        elif v < p25:
            out.append("highly oversold")
        elif v < p50:
            out.append("slightly oversold")
        else:
            out.append("neutral")
    return out


def _classify_cross(fast, slow, up_label, down_label, neutral="neutral") -> List:
    out = []
    for f, s in zip(_finite(fast), _finite(slow)):
        if pd.isna(f) or pd.isna(s):
            out.append("")
        elif f > s:
            out.append(up_label)
        elif f < s:
            out.append(down_label)
        else:
            out.append(neutral)
    return out


def bollinger_decision_states(
    close: pd.Series, period: int = 20, std_dev: float = 2.0
) -> Tuple[List, List, List, List]:
    close = _finite(close)
    upper, middle, lower = compute_bollinger_bands(close, period, std_dev)
    width = upper - lower
    pct_b = (close - lower) / width.where(width != 0, np.nan)
    bandwidth = width / middle.where(middle != 0, np.nan)
    bw_finite = _finite(bandwidth).dropna()
    bw_low = bw_high = None
    if len(bw_finite) >= RSI_BIN_MIN_HISTORY:
        bw_low = float(bw_finite.quantile(0.25))
        bw_high = float(bw_finite.quantile(0.75))

    indicator, trend, volatility, obos = [], [], [], []
    for pb, bw in zip(pct_b, _finite(bandwidth)):
        t = o = ""
        if not pd.isna(pb):
            o = "overbought" if pb >= 1.0 else ("oversold" if pb <= 0.0 else "neutral")
            t = "bullish" if pb > 0.6 else ("bearish" if pb < 0.4 else "neutral")
        v = ""
        if not pd.isna(bw):
            if bw_low is None:
                v = "normal volatility"
            elif bw > bw_high:
                v = "high volatility"
            elif bw < bw_low:
                v = "low volatility"
            else:
                v = "normal volatility"
        trend.append(t)
        volatility.append(v)
        obos.append(o)
        indicator.append(" | ".join(p for p in (t, o, v) if p))
    return indicator, trend, volatility, obos


def compute_decision_features(
    df: pd.DataFrame,
    rsi7_thresholds=None,
    rsi14_thresholds=None,
) -> pd.DataFrame:
    out = pd.DataFrame(
        index=df.index, columns=DECISION_FEATURE_COLUMNS, dtype=object
    )
    close_col = find_close_column(df)
    if close_col is None:
        return out.fillna("")
    try:
        close = _finite(df[close_col])
        close.index = df.index
        rsi7 = compute_rsi(close, 7)
        rsi14 = compute_rsi(close, 14)
        out["RSI_7"] = [
            "" if pd.isna(v) else round(float(v), 2) for v in _finite(rsi7)
        ]
        out["SMA_5_20"] = _classify_cross(
            compute_sma(close, 5), compute_sma(close, 20), "bullish", "bearish"
        )
        out["SMA_20_50"] = _classify_cross(
            compute_sma(close, 20), compute_sma(close, 50), "strong bullish", "strong bearish"
        )
        out["EMA_12_26"] = _classify_cross(
            compute_ema(close, 12), compute_ema(close, 26), "bullish", "bearish"
        )
        out["EMA_26_50"] = _classify_cross(
            compute_ema(close, 26), compute_ema(close, 50), "long-term bullish", "long-term bearish"
        )
        bb_ind, bb_trend, bb_vol, bb_obos = bollinger_decision_states(close)
        out["BB_indicator"] = bb_ind
        out["BB_trend_state"] = bb_trend
        out["BB_volatility_state"] = bb_vol
        out["BB_overbought_oversold"] = bb_obos
        out["RSI_indicator_7"] = classify_rsi_bins(rsi7, rsi7_thresholds)
        out["RSI_indicator_14"] = classify_rsi_bins(rsi14, rsi14_thresholds)
    except Exception as exc:
        _log.warning("Decision feature computation failed: %s", exc)
    return out.fillna("")


def compute_decision_rsi_history(
    df: pd.DataFrame, periods: Tuple[int, ...] = (7, 14)
) -> Dict[int, List[float]]:
    history: Dict[int, List[float]] = {p: [] for p in periods}
    close_col = find_close_column(df)
    if close_col is None:
        return history
    close = _finite(df[close_col])
    for period in periods:
        try:
            history[period] = _finite(compute_rsi(close, period)).dropna().tolist()
        except Exception:
            history[period] = []
    return history


# ===========================================================================
# Main feature engineering function
# ===========================================================================

def compute_indicators(df: pd.DataFrame) -> pd.DataFrame:
    """Compute 24 technical indicators + 30 forward-return labels for one symbol.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame containing at minimum Open, High, Low, Close, Volume
        columns (case-insensitive, prefix-tolerant).  Extra columns (Date,
        Date_str, __sheet_row_number, __sort_position, etc.) are preserved
        unchanged.

    Returns
    -------
    pd.DataFrame
        The original DataFrame with indicator and label columns appended.
        All indicators are lagged by one step (``shift(1)``); labels are not.

    Raises
    ------
    FeatureEngineeringError
        If OHLCV columns cannot be located, or if no valid OHLCV rows remain
        after NaN filtering.  Callers that previously relied on the silent
        skip-and-return behaviour should catch this exception.
    """
    t0 = time.monotonic()
    df_work = df.copy()

    # ── Step 1: locate OHLCV ────────────────────────────────────────────────
    ohlcv_map = find_ohlcv_columns(df_work)
    if ohlcv_map is None:
        available = list(df_work.columns)
        raise FeatureEngineeringError(
            f"Cannot locate OHLCV columns. "
            f"Available columns: {available}. "
            f"Expected lowercase or Title-Case: open/Open, high/High, low/Low, "
            f"close/Close, volume/Volume (or Open_SYMBOL prefixed variants)."
        )

    _log.debug(
        "[FE] OHLCV located — open=%s high=%s low=%s close=%s volume=%s",
        ohlcv_map["open"], ohlcv_map["high"], ohlcv_map["low"],
        ohlcv_map["close"], ohlcv_map["volume"],
    )

    # ── Step 2: coerce to numeric and filter NaN rows ───────────────────────
    open_s  = pd.to_numeric(df_work[ohlcv_map["open"]],   errors="coerce")
    high_s  = pd.to_numeric(df_work[ohlcv_map["high"]],   errors="coerce")
    low_s   = pd.to_numeric(df_work[ohlcv_map["low"]],    errors="coerce")
    close_s = pd.to_numeric(df_work[ohlcv_map["close"]],  errors="coerce")
    vol_s   = pd.to_numeric(df_work[ohlcv_map["volume"]], errors="coerce")

    valid_mask = ~(
        open_s.isna() | high_s.isna() | low_s.isna() | close_s.isna() | vol_s.isna()
    )
    if not valid_mask.any():
        raise FeatureEngineeringError(
            "No valid OHLCV rows remain after NaN filtering. "
            "The input DataFrame has no rows where all five price columns are finite."
        )

    dropped = int((~valid_mask).sum())
    if dropped > 0:
        _log.warning("[FE] Dropped %d rows with NaN OHLCV values", dropped)

    df_work = df_work.loc[valid_mask].reset_index(drop=True)
    open_s  = open_s.loc[valid_mask].reset_index(drop=True)
    high_s  = high_s.loc[valid_mask].reset_index(drop=True)
    low_s   = low_s.loc[valid_mask].reset_index(drop=True)
    close_s = close_s.loc[valid_mask].reset_index(drop=True)
    vol_s   = vol_s.loc[valid_mask].reset_index(drop=True)

    _log.debug("[FE] Processing %d valid OHLCV rows", len(df_work))

    # ── Step 3: compute indicators ──────────────────────────────────────────
    indicators: Dict[str, pd.Series] = {}

    try:
        indicators["RSI_14"] = compute_rsi(close_s, 14)
    except Exception as exc:
        _log.warning("[FE] RSI_14 failed: %s", exc)
        indicators["RSI_14"] = pd.Series(np.nan, index=range(len(df_work)))

    try:
        macd, sig, hist = compute_macd(close_s, 12, 26, 9)
        indicators["MACD_12_26"]    = macd.reset_index(drop=True)
        indicators["MACD_Signal_9"] = sig.reset_index(drop=True)
        indicators["MACD_Histogram"]= hist.reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] MACD failed: %s", exc)

    try:
        k, d = compute_stochastic(high_s, low_s, close_s, 14, 3)
        indicators["Stochastic_%K"] = k.reset_index(drop=True)
        indicators["Stochastic_%D"] = d.reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] Stochastic failed: %s", exc)

    try:
        indicators["SMA_5"]  = compute_sma(close_s, 5).reset_index(drop=True)
        indicators["SMA_20"] = compute_sma(close_s, 20).reset_index(drop=True)
        indicators["SMA_50"] = compute_sma(close_s, 50).reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] SMA failed: %s", exc)

    try:
        indicators["EMA_12"] = compute_ema(close_s, 12).reset_index(drop=True)
        indicators["EMA_26"] = compute_ema(close_s, 26).reset_index(drop=True)
        indicators["EMA_50"] = compute_ema(close_s, 50).reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] EMA failed: %s", exc)

    try:
        indicators["ADX_14"] = compute_adx(high_s, low_s, close_s, 14).reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] ADX_14 failed: %s", exc)

    try:
        ub, mb, lb = compute_bollinger_bands(close_s, 20, 2)
        indicators["BB_Upper_20"]  = ub.reset_index(drop=True)
        indicators["BB_Middle_20"] = mb.reset_index(drop=True)
        indicators["BB_Lower_20"]  = lb.reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] Bollinger Bands failed: %s", exc)

    try:
        indicators["ATR_14"] = compute_atr(high_s, low_s, close_s, 14).reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] ATR_14 failed: %s", exc)

    try:
        indicators["OBV"] = compute_obv(close_s, vol_s).reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] OBV failed: %s", exc)

    try:
        indicators["VWAP"] = compute_vwap(high_s, low_s, close_s, vol_s).reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] VWAP failed: %s", exc)

    try:
        indicators["Daily_Return_%"] = compute_daily_return(close_s).reset_index(drop=True)
        indicators["Log_Return_%"]   = compute_log_return(close_s).reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] Return indicators failed: %s", exc)

    try:
        indicators["ROC_12"] = compute_roc(close_s, 12).reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] ROC_12 failed: %s", exc)

    try:
        indicators["CCI_20"] = compute_cci(high_s, low_s, close_s, 20).reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] CCI_20 failed: %s", exc)

    try:
        indicators["Williams_%R"] = compute_williams_r(high_s, low_s, close_s, 14).reset_index(drop=True)
    except Exception as exc:
        _log.warning("[FE] Williams_%R failed: %s", exc)

    # ── Step 4: attach to df_work (lag-1 to prevent look-ahead) ────────────
    for col_name, series in indicators.items():
        if isinstance(series, pd.Series):
            df_work[col_name] = series.shift(1).values
        else:
            df_work[col_name] = np.nan

    indicators_computed = len(indicators)
    _log.debug("[FE] %d indicator columns computed and lagged by 1 step", indicators_computed)

    # ── Step 5: forward-return labels (P0.1) — NOT shifted ─────────────────
    try:
        close_for_labels = pd.to_numeric(
            df_work[ohlcv_map["close"]], errors="coerce"
        ).reset_index(drop=True)
        forward_labels = compute_forward_log_returns(close_for_labels)
        for label_name, label_series in forward_labels.items():
            df_work[label_name] = label_series.values
        label_cols = list(forward_labels.keys())
        df_work[HAS_LABELS_COL] = df_work[label_cols].notna().all(axis=1).values
        labels_with_data = int(df_work[HAS_LABELS_COL].sum())
        _log.debug(
            "[FE] %d forward-return label columns attached; %d rows have complete labels",
            len(forward_labels), labels_with_data,
        )
    except Exception as exc:
        _log.warning("[FE] Forward label attachment failed: %s", exc)
        for label_name in forward_label_columns():
            df_work[label_name] = np.nan
        df_work[HAS_LABELS_COL] = False

    elapsed = time.monotonic() - t0
    _log.info(
        "[FE] compute_indicators complete — rows=%d indicators=%d elapsed=%.3fs",
        len(df_work), indicators_computed, elapsed,
    )
    return df_work


# ===========================================================================
# Workbook I/O helpers (standalone batch mode only)
# ===========================================================================

def load_workbook_sheets(filepath: str) -> Dict[str, pd.DataFrame]:
    """Load all sheets from an Excel workbook."""
    try:
        sheets = pd.read_excel(filepath, sheet_name=None)
        _log.info("[FE:batch] Loaded %d sheets from %s", len(sheets), filepath)
        return sheets
    except FileNotFoundError:
        _log.error("[FE:batch] Workbook not found: %s", filepath)
        raise
    except Exception as exc:
        _log.error("[FE:batch] Error loading workbook %s: %s", filepath, exc)
        raise


def write_updated_workbook(
    sheets_dict: Dict[str, pd.DataFrame], output_filepath: str
) -> None:
    """Write updated sheets back to the Excel workbook atomically.

    Uses a temp-file-then-rename strategy to avoid partial writes that corrupt
    the workbook if the process is interrupted mid-write.
    """
    output_path = Path(output_filepath)
    tmp_path = output_path.with_suffix(".tmp.xlsx")
    try:
        with pd.ExcelWriter(str(tmp_path), engine="openpyxl") as writer:
            for sheet_name, df in sheets_dict.items():
                df_clean = df.dropna(how="all")
                df_clean.to_excel(writer, sheet_name=sheet_name, index=False)
        # Atomic replace
        tmp_path.replace(output_path)
        _log.info("[FE:batch] Workbook written atomically to %s", output_filepath)
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        _log.error("[FE:batch] Failed to write workbook: %s", exc)
        raise


# ===========================================================================
# Standalone batch entry point (offline training-data refresh only)
# ===========================================================================

def main() -> None:
    """Offline batch mode: read training workbook, engineer features, write back.

    This function is NOT part of the Docker workflow.  It is intended for
    offline use when refreshing the training corpus (Data/nse_stock_data_train.xlsx).

    Usage::

        python Feature_Engineering.py [--workbook Data/nse_stock_data_train.xlsx]

    """
    import argparse
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    parser = argparse.ArgumentParser(
        description="Offline feature engineering: read OHLCV workbook, compute indicators, write back."
    )
    parser.add_argument(
        "--workbook",
        default=None,
        help=(
            "Path to the Excel workbook to process. "
            "Defaults to Data/nse_stock_data_train.xlsx if it exists, "
            "otherwise Data/nse_stock_data.xlsx."
        ),
    )
    args = parser.parse_args()

    script_dir = Path(__file__).parent
    data_dir   = script_dir / "Data"

    # Determine input workbook — prefer the training corpus that actually exists
    if args.workbook:
        input_file = Path(args.workbook)
    else:
        candidate_train = data_dir / "nse_stock_data_train.xlsx"
        candidate_std   = data_dir / "nse_stock_data.xlsx"
        if candidate_train.exists():
            input_file = candidate_train
        elif candidate_std.exists():
            input_file = candidate_std
        else:
            _log.error(
                "[FE:batch] No workbook found. Tried:\n  %s\n  %s\n"
                "Pass --workbook <path> explicitly.",
                candidate_train, candidate_std,
            )
            sys.exit(1)

    output_file = input_file  # overwrite in place (atomic write is safe)

    if not input_file.exists():
        _log.error("[FE:batch] Workbook not found: %s", input_file)
        sys.exit(1)

    _log.info("=" * 70)
    _log.info("NSE Stock Data — Offline Feature Engineering (batch mode)")
    _log.info("=" * 70)
    _log.info("[FE:batch] Input  : %s  (%s bytes)", input_file, f"{input_file.stat().st_size:,}")
    _log.info("[FE:batch] Output : %s", output_file)

    sheets = load_workbook_sheets(str(input_file))
    if not sheets:
        _log.error("[FE:batch] Workbook contains no sheets.")
        sys.exit(1)

    processed: Dict[str, pd.DataFrame] = {}
    skipped = 0
    total = len(sheets)
    t_start = time.monotonic()

    for sheet_name, df in sheets.items():
        _log.info("[FE:batch] Processing sheet '%s' (%d rows × %d cols)", sheet_name, len(df), len(df.columns))

        if df.empty:
            _log.warning("[FE:batch]   Sheet is empty — skipping")
            processed[sheet_name] = df
            skipped += 1
            continue

        # Date detection and cleanup
        date_col = detect_date_column(df)
        if date_col is None:
            _log.warning("[FE:batch]   No date column detected — skipping indicators")
            processed[sheet_name] = df
            skipped += 1
            continue

        _log.info("[FE:batch]   Date column: '%s'", date_col)
        df = ensure_date_column(df, date_col)

        try:
            df = compute_indicators(df)
            _log.info(
                "[FE:batch]   Done: %d rows, %d columns (including %d indicators, %d labels)",
                len(df),
                len(df.columns),
                len(INDICATOR_COLUMNS),
                len(FORWARD_HORIZONS),
            )
        except FeatureEngineeringError as exc:
            _log.warning("[FE:batch]   Skipped: %s", exc)
            skipped += 1

        processed[sheet_name] = df

    elapsed = time.monotonic() - t_start
    _log.info("-" * 70)
    _log.info(
        "[FE:batch] Processed %d/%d sheets (%d skipped) in %.1fs",
        total - skipped, total, skipped, elapsed,
    )

    write_updated_workbook(processed, str(output_file))

    _log.info("=" * 70)
    _log.info("[FE:batch] Feature engineering complete.")
    _log.info("=" * 70)


if __name__ == "__main__":
    main()
